import customtkinter
from tkinter import *
import openpyxl as xl
from datetime import date
from PIL import Image
from tkinter import messagebox

app = customtkinter.CTk()
app.geometry("800x450")
app.config(bg="#000000")
app.title("Supermarket")

img1=customtkinter.CTkImage(Image.open(r"chocolate.png"))
img2=customtkinter.CTkImage(Image.open(r"juice.png"))
img3=customtkinter.CTkImage(Image.open(r"soda.png"))
img4=customtkinter.CTkImage(Image.open(r"apple.png"))

font1=('Arial', 15, 'bold')

Variable1=IntVar()
Variable2=IntVar()
Variable3=IntVar()
Variable4=IntVar()
Variable5=0
Variable6=''

price_list=[15.000,10.000,8.000,3.000]

def gajadi():
    Variable1.set(0)
    Variable2.set(0)
    Variable3.set(0)
    Variable4.set(0)
    name_entry.delete(0, 'end')
    total_label=customtkinter.CTkLabel(app,text=Variable5,font=font1,text_color="#000000",bg_color="#000000",width=300,justify=CENTER)
    total_label.place(x=450,y=148)
    current_date_label=customtkinter.CTkLabel(app,text=Variable6,font=font1,text_color="#000000",bg_color="#000000",width=300)
    current_date_label.place(x=450,y=220)

def bayar():
    global Variable5,Variable6
    if(name_entry.get()==''):
        messagebox.showerror(title="error",message="Masukkan Nama Dulu")
    else:
        Variable5=Variable1.get()*price_list[0]+Variable2.get()*price_list[1]+Variable3.get()*price_list[2]+Variable4.get()*price_list[2]
        total_label=customtkinter.CTkLabel(app,text=Variable5,font=font1,text_color="#000000",bg_color="#FFFFFF",width=300,justify=CENTER)
        total_label.place(x=450,y=148)
        Variable6=date.today()
        current_date_label=customtkinter.CTkLabel(app,text=Variable6,font=font1,text_color="#000000",bg_color="#FFFFFF",width=300)
        current_date_label.place(x=450,y=220)

def gas():
    global Variable5,Variable6
    file=xl.load_workbook('pembeli.xlsx')
    sheet=file["Sheet1"]
    sheet.cell(column=1,row=sheet.max_row+1,value=name_entry.get())
    sheet.cell(column=2,row=sheet.max_row,value=Variable5)
    sheet.cell(column=3,row=sheet.max_row,value=Variable6)
    file.save('pembeli.xlsx')
    messagebox.showinfo(title="Berhasil",message="Data Tersimpan.")

button1 = customtkinter.CTkButton(app,text="Cokelat",fg_color='#0e1d54',font=font1,hover_color='#0e1d54',text_color="#FFFFFF",bg_color="#000000",border_width=1,width=100,height=100,image=img1,compound=TOP)
button1.place(x=30,y=20)

button2 = customtkinter.CTkButton(app,text="Jus",fg_color='#0e1d54',font=font1,hover_color='#0e1d54',text_color="#FFFFFF",bg_color="#000000",border_width=1,width=100,height=100,image=img2,compound=TOP)
button2.place(x=300,y=20)

button3 = customtkinter.CTkButton(app,text="Soda",fg_color='#0e1d54',font=font1,hover_color='#0e1d54',text_color="#FFFFFF",bg_color="#000000",border_width=1,width=100,height=100,image=img3,compound=TOP)
button3.place(x=30,y=240)

button4 = customtkinter.CTkButton(app,text="Apel",fg_color='#0e1d54',font=font1,hover_color='#0e1d54',text_color="#FFFFFF",bg_color="#000000",border_width=1,width=100,height=100,image=img4,compound=TOP)
button4.place(x=300,y=240)

sp1=Spinbox(app,from_=0,to=10,font=font1,textvariable=Variable1,width=8,background="#5c99ad",justify=CENTER)
sp1.place(x=50,y=160)

sp2=Spinbox(app,from_=0,to=10,font=font1,textvariable=Variable2,width=8,background="#5c99ad",justify=CENTER)
sp2.place(x=385,y=160)

sp3=Spinbox(app,from_=0,to=10,font=font1,textvariable=Variable3,width=8,background="#5c99ad",justify=CENTER)
sp3.place(x=50,y=435)

sp4=Spinbox(app,from_=0,to=10,font=font1,textvariable=Variable4,width=8,background="#5c99ad",justify=CENTER)
sp4.place(x=385,y=435)

name_label=customtkinter.CTkLabel(app,text="Nama Pembeli:", font=font1, text_color="#FFFFFF",bg_color="#000000")
name_label.place(x=450,y=50)

name_entry=customtkinter.CTkEntry(app,font=font1,text_color="#000000",bg_color="#000000",border_color="#000000",width=300)
name_entry.place(x=450,y=80)

price_label=customtkinter.CTkLabel(app,text="Total Harga:", font=font1, text_color="#FFFFFF",bg_color="#000000")
price_label.place(x=450,y=120)

date_label=customtkinter.CTkLabel(app,text="Tanggal Pembelian:", font=font1, text_color="#FFFFFF",bg_color="#000000")
date_label.place(x=450,y=190)

pay_button=customtkinter.CTkButton(app,command=bayar,text="Bayar",font=font1,bg_color="#000000",hover_color="#43158a")
pay_button.place(x=450,y=280)

cancel_button=customtkinter.CTkButton(app,command=gajadi,text="Gajadi",font=font1,bg_color="#000000",hover_color="#43158a")
cancel_button.place(x=600,y=280)

confirm_button=customtkinter.CTkButton(app,command=gas,text="Gas",font=font1,bg_color="#000000",hover_color="#43158a")
confirm_button.place(x=525,y=320)

app.mainloop()